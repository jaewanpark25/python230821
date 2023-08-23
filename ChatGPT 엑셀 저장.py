#ChatGPT 엑셀 저장.py

import openpyxl
import random
from datetime import datetime, timedelta

# Create a new workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write headers
headers = ["Product Name", "Price ($)", "Quantity", "Sale Date"]
for col_idx, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col_idx, value=header)

# Generate and write sales data
product_names = ["Laptop", "Smartphone", "Tablet", "Headphones", "Monitor"]
start_date = datetime(2023, 1, 1)
for row_idx in range(2, 102):
    product_name = random.choice(product_names)
    price = round(random.uniform(200, 2000), 2)
    quantity = random.randint(1, 10)
    sale_date = start_date + timedelta(days=random.randint(1, 365))

    sheet.cell(row=row_idx, column=1, value=product_name)
    sheet.cell(row=row_idx, column=2, value=price)
    sheet.cell(row=row_idx, column=3, value=quantity)
    sheet.cell(row=row_idx, column=4, value=sale_date)

# Save the workbook to the specified path
file_path = "c:\\work\\sales.xlsx"
workbook.save(file_path)
print(f"Sales data saved to {file_path}")


